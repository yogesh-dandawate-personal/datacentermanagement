"""
Excel Report Generation Service

Generates Excel workbooks with:
- Summary sheet with key metrics
- Detailed data sheets (energy, carbon, KPI)
- Embedded charts
- Evidence links sheet
- Formatted cells with colors, fonts, borders
- Frozen header rows
- Auto-fitted column widths
"""

from io import BytesIO
from datetime import datetime
from typing import Dict, List, Optional, Any
from decimal import Decimal
import logging
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment,
    numbers
)
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models import (
    Report, Organization, KPIDefinition, KPISnapshot,
    ComplianceReport, ComplianceTarget
)

logger = logging.getLogger(__name__)


class ExcelGenerator:
    """Generate Excel reports with formatting and charts"""

    def __init__(self, db: Session):
        """
        Initialize Excel Generator

        Args:
            db: SQLAlchemy database session
        """
        self.db = db
        self.workbook = None
        self._setup_styles()

    def _setup_styles(self) -> None:
        """Setup reusable cell styles"""
        self.header_fill = PatternFill(
            start_color="2B6CB0",
            end_color="2B6CB0",
            fill_type="solid"
        )
        self.header_font = Font(
            bold=True,
            color="FFFFFF",
            size=11
        )

        self.subheader_fill = PatternFill(
            start_color="D9E8F5",
            end_color="D9E8F5",
            fill_type="solid"
        )
        self.subheader_font = Font(bold=True, size=10)

        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        self.thin_border = Border(
            left=Side(style='thin', color="CCCCCC"),
            right=Side(style='thin', color="CCCCCC"),
            top=Side(style='thin', color="CCCCCC"),
            bottom=Side(style='thin', color="CCCCCC")
        )

        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )

    def _format_header_row(self, worksheet, row_num: int, num_cols: int) -> None:
        """
        Format header row with styling

        Args:
            worksheet: Openpyxl worksheet object
            row_num: Row number (1-indexed)
            num_cols: Number of columns to format
        """
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border

    def _auto_fit_columns(self, worksheet, max_col: int) -> None:
        """
        Auto-fit column widths

        Args:
            worksheet: Openpyxl worksheet object
            max_col: Number of columns
        """
        for col in range(1, max_col + 1):
            column_letter = get_column_letter(col)
            max_length = 0

            for row in worksheet.iter_rows(min_col=col, max_col=col):
                for cell in row:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

    def _create_summary_sheet(
        self,
        organization_name: str,
        period_start: datetime,
        period_end: datetime,
        scope_1: Optional[Decimal] = None,
        scope_2: Optional[Decimal] = None,
        scope_3: Optional[Decimal] = None,
        key_metrics: Optional[Dict] = None
    ) -> None:
        """
        Create summary sheet with key metrics

        Args:
            organization_name: Organization name
            period_start: Report period start
            period_end: Report period end
            scope_1: Scope 1 emissions
            scope_2: Scope 2 emissions
            scope_3: Scope 3 emissions
            key_metrics: Dictionary of key metrics
        """
        ws = self.workbook.active
        ws.title = "Summary"

        # Header
        ws['A1'] = "Report Summary"
        ws['A1'].font = Font(bold=True, size=14, color="1a365d")
        ws.merge_cells('A1:C1')

        # Organization and period
        ws['A3'] = "Organization:"
        ws['B3'] = organization_name
        ws['A4'] = "Report Period:"
        ws['B4'] = f"{period_start.strftime('%Y-%m-%d')} to {period_end.strftime('%Y-%m-%d')}"
        ws['A5'] = "Report Date:"
        ws['B5'] = datetime.utcnow().strftime('%Y-%m-%d')

        # Key metrics table
        ws['A7'] = "Metric"
        ws['B7'] = "Value"
        ws['C7'] = "Unit"
        self._format_header_row(ws, 7, 3)

        row = 8
        metrics = [
            ("Scope 1 Emissions", f"{float(scope_1) if scope_1 else 0:.2f}", "Tonnes CO2e"),
            ("Scope 2 Emissions", f"{float(scope_2) if scope_2 else 0:.2f}", "Tonnes CO2e"),
            ("Scope 3 Emissions", f"{float(scope_3) if scope_3 else 0:.2f}", "Tonnes CO2e"),
            ("Total Emissions", f"{float((scope_1 or 0) + (scope_2 or 0) + (scope_3 or 0)):.2f}", "Tonnes CO2e"),
        ]

        for metric, value, unit in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=unit)

            ws.cell(row=row, column=2).number_format = '#,##0.00'
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = self.thin_border
            row += 1

        # Additional metrics
        if key_metrics:
            row += 1
            for metric_name, metric_value in key_metrics.items():
                if metric_name not in ['scope_1', 'scope_2', 'scope_3']:
                    ws.cell(row=row, column=1, value=metric_name.replace('_', ' ').title())
                    value_cell = ws.cell(row=row, column=2, value=metric_value)
                    if isinstance(metric_value, (int, float, Decimal)):
                        value_cell.number_format = '#,##0.00'
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).border = self.thin_border
                    row += 1

        # Freeze panes
        ws.freeze_panes = "A8"

        # Auto-fit columns
        self._auto_fit_columns(ws, 3)

    def _create_detailed_sheet(
        self,
        sheet_name: str,
        data: List[List[Any]],
        headers: List[str]
    ) -> None:
        """
        Create detailed data sheet

        Args:
            sheet_name: Name of worksheet
            data: List of data rows
            headers: Column headers
        """
        ws = self.workbook.create_sheet(sheet_name)

        # Add headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border

        # Add data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value

                # Format numeric cells
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'

                cell.border = self.thin_border
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color="F5F5F5",
                        end_color="F5F5F5",
                        fill_type="solid"
                    )

        # Freeze header
        ws.freeze_panes = "A2"

        # Auto-fit columns
        self._auto_fit_columns(ws, len(headers))

    def _create_kpi_sheet(self, organization_id: UUID) -> None:
        """
        Create KPI performance sheet

        Args:
            organization_id: Organization UUID
        """
        try:
            kpis = self.db.query(KPIDefinition).filter(
                KPIDefinition.organization_id == organization_id,
                KPIDefinition.is_active == True
            ).all()

            ws = self.workbook.create_sheet("KPIs")

            headers = ["KPI Name", "Latest Value", "Target Value", "Unit", "Status", "Variance %"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = self.center_alignment
                cell.border = self.border

            row = 2
            for kpi in kpis:
                latest_snapshot = self.db.query(KPISnapshot).filter(
                    KPISnapshot.kpi_id == kpi.id
                ).order_by(KPISnapshot.snapshot_date.desc()).first()

                if latest_snapshot:
                    ws.cell(row=row, column=1, value=kpi.kpi_name)
                    ws.cell(row=row, column=2, value=float(latest_snapshot.calculated_value))
                    ws.cell(row=row, column=3, value=float(latest_snapshot.target_value) if latest_snapshot.target_value else 0)
                    ws.cell(row=row, column=4, value=kpi.unit or "")
                    ws.cell(row=row, column=5, value=latest_snapshot.status or "normal")
                    ws.cell(row=row, column=6, value=float(latest_snapshot.variance_percent) if latest_snapshot.variance_percent else 0)

                    for col in range(1, 7):
                        cell = ws.cell(row=row, column=col)
                        if col in [2, 3, 6]:
                            cell.number_format = '#,##0.00'
                        cell.border = self.thin_border

                    row += 1

            ws.freeze_panes = "A2"
            self._auto_fit_columns(ws, 6)

        except Exception as e:
            logger.warning(f"Error creating KPI sheet: {str(e)}")

    def _create_evidence_sheet(self, evidence_links: Optional[List[Dict]] = None) -> None:
        """
        Create evidence references sheet

        Args:
            evidence_links: List of evidence references
        """
        ws = self.workbook.create_sheet("Evidence")

        headers = ["Document Name", "Upload Date", "Type", "URL"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_alignment
            cell.border = self.border

        if evidence_links:
            for row_idx, evidence in enumerate(evidence_links, 2):
                ws.cell(row=row_idx, column=1, value=evidence.get('name', 'N/A'))
                ws.cell(row=row_idx, column=2, value=evidence.get('uploaded_at', 'N/A'))
                ws.cell(row=row_idx, column=3, value=evidence.get('type', 'N/A'))
                ws.cell(row=row_idx, column=4, value=evidence.get('link', 'N/A'))

                for col in range(1, 5):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.border = self.thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(
                            start_color="F5F5F5",
                            end_color="F5F5F5",
                            fill_type="solid"
                        )

        ws.freeze_panes = "A2"
        self._auto_fit_columns(ws, 4)

    def generate_workbook(
        self,
        report_id: UUID,
        organization_id: UUID,
        scope_1: Optional[Decimal] = None,
        scope_2: Optional[Decimal] = None,
        scope_3: Optional[Decimal] = None,
        key_metrics: Optional[Dict] = None,
        evidence_links: Optional[List[Dict]] = None,
        detailed_data: Optional[Dict[str, List[Any]]] = None
    ) -> BytesIO:
        """
        Generate complete Excel workbook

        Args:
            report_id: Report UUID
            organization_id: Organization UUID
            scope_1: Scope 1 emissions
            scope_2: Scope 2 emissions
            scope_3: Scope 3 emissions
            key_metrics: Dictionary of key metrics
            evidence_links: List of evidence references
            detailed_data: Dictionary of sheet_name -> data

        Returns:
            BytesIO object containing Excel workbook

        Raises:
            ValueError: If report or organization not found
        """
        report = self.db.query(Report).filter(Report.id == report_id).first()
        if not report:
            raise ValueError(f"Report {report_id} not found")

        organization = self.db.query(Organization).filter(
            Organization.id == organization_id
        ).first()
        if not organization:
            raise ValueError(f"Organization {organization_id} not found")

        # Create workbook
        self.workbook = Workbook()

        # Create summary sheet
        self._create_summary_sheet(
            organization.name,
            report.report_period_start,
            report.report_period_end,
            scope_1, scope_2, scope_3,
            key_metrics
        )

        # Create KPI sheet
        self._create_kpi_sheet(organization_id)

        # Create detailed data sheets
        if detailed_data:
            for sheet_name, (headers, data) in detailed_data.items():
                self._create_detailed_sheet(sheet_name, data, headers)

        # Create evidence sheet
        self._create_evidence_sheet(evidence_links)

        # Save to BytesIO
        excel_buffer = BytesIO()
        try:
            self.workbook.save(excel_buffer)
            excel_buffer.seek(0)
            logger.info(f"Excel workbook generated successfully for report {report_id}")
            return excel_buffer
        except Exception as e:
            logger.error(f"Error generating Excel workbook for report {report_id}: {str(e)}")
            raise
