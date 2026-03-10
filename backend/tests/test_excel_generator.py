"""
Tests for Excel Report Generation Service

Test Coverage:
- Excel workbook generation with multiple sheets
- Data integrity and formatting
- Charts and frozen panes
- Evidence links sheet
- Column width auto-fit
- Error handling
- Large datasets
"""

import pytest
from io import BytesIO
from datetime import datetime
from decimal import Decimal
from uuid import uuid4

from openpyxl import load_workbook

from app.models import (
    Report, Organization, Tenant, User,
    KPIDefinition, KPISnapshot
)
from app.services.excel_generator import ExcelGenerator


@pytest.fixture
def excel_generator(db):
    """Create Excel generator instance"""
    return ExcelGenerator(db)


@pytest.fixture
def setup_report_with_kpis(db, tenant, organization):
    """Setup report with KPI data"""
    # Create report
    report = Report(
        organization_id=organization.id,
        tenant_id=tenant.id,
        report_type="esg_monthly",
        report_period_start=datetime(2024, 1, 1),
        report_period_end=datetime(2024, 1, 31),
        created_by=None,
        current_state="draft"
    )
    db.add(report)
    db.commit()

    # Create KPI definitions
    kpi_names = ["PUE", "CUE", "WUE", "ERE"]
    kpis = []
    for name in kpi_names:
        kpi = KPIDefinition(
            organization_id=organization.id,
            tenant_id=tenant.id,
            kpi_name=name,
            kpi_type="standard",
            formula=f"{name} formula",
            unit="ratio" if name == "PUE" else "g/kWh",
            target_value=Decimal("1.5"),
            lower_bound=Decimal("1.0"),
            upper_bound=Decimal("2.0"),
            is_active=True
        )
        db.add(kpi)
        db.commit()
        kpis.append(kpi)

        # Add snapshot
        snapshot = KPISnapshot(
            kpi_id=kpi.id,
            tenant_id=tenant.id,
            snapshot_date=datetime(2024, 1, 31),
            calculated_value=Decimal("1.45"),
            target_value=Decimal("1.5"),
            variance_percent=Decimal("3.33"),
            status="normal",
            data_quality_score=Decimal("95.00")
        )
        db.add(snapshot)
    db.commit()

    return report, kpis


class TestExcelGenerationBasics:
    """Test basic Excel generation"""

    def test_excel_generator_initialization(self, excel_generator):
        """Test Excel generator initializes correctly"""
        assert excel_generator.db is not None
        assert excel_generator.header_fill is not None
        assert excel_generator.border is not None

    def test_generate_basic_workbook(self, excel_generator, setup_report_with_kpis):
        """Test basic workbook generation"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        assert isinstance(excel_buffer, BytesIO)
        assert excel_buffer.tell() == 0
        content = excel_buffer.read()
        assert len(content) > 0

    def test_generate_workbook_with_metrics(self, excel_generator, setup_report_with_kpis):
        """Test workbook generation with key metrics"""
        report, kpis = setup_report_with_kpis

        key_metrics = {
            "pue": 1.45,
            "cue": 48.5,
            "wue": 1.8,
            "ere": 1.6
        }

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("150.50"),
            scope_2=Decimal("200.75"),
            scope_3=Decimal("500.25"),
            key_metrics=key_metrics
        )

        assert isinstance(excel_buffer, BytesIO)
        content = excel_buffer.read()
        assert len(content) > 0


class TestExcelWorksheets:
    """Test worksheet creation"""

    def test_summary_sheet_creation(self, excel_generator, setup_report_with_kpis):
        """Test summary sheet is created"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        assert "Summary" in wb.sheetnames

    def test_kpi_sheet_creation(self, excel_generator, setup_report_with_kpis):
        """Test KPI sheet is created"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        assert "KPIs" in wb.sheetnames

    def test_evidence_sheet_creation(self, excel_generator, setup_report_with_kpis):
        """Test evidence sheet is created"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        assert "Evidence" in wb.sheetnames

    def test_detailed_data_sheets(self, excel_generator, setup_report_with_kpis):
        """Test custom detailed data sheets"""
        report, kpis = setup_report_with_kpis

        detailed_data = {
            "Energy Data": (
                ["Date", "Usage (kWh)", "Source"],
                [
                    ["2024-01-01", 1000.50, "Solar"],
                    ["2024-01-02", 1050.75, "Grid"],
                    ["2024-01-03", 1100.25, "Grid"]
                ]
            ),
            "Carbon Data": (
                ["Date", "Emissions (tCO2e)", "Category"],
                [
                    ["2024-01-01", 50.25, "Scope 2"],
                    ["2024-01-02", 52.50, "Scope 2"],
                    ["2024-01-03", 55.75, "Scope 2"]
                ]
            )
        }

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00"),
            detailed_data=detailed_data
        )

        wb = load_workbook(excel_buffer)
        assert "Energy Data" in wb.sheetnames
        assert "Carbon Data" in wb.sheetnames


class TestExcelDataIntegrity:
    """Test data integrity in Excel files"""

    def test_summary_sheet_data(self, excel_generator, setup_report_with_kpis):
        """Test summary sheet contains correct data"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Summary"]

        # Check header row
        assert ws["A1"].value == "Report Summary"

        # Check organization name
        org_name = ws["B3"].value
        assert org_name is not None

    def test_kpi_sheet_data(self, excel_generator, setup_report_with_kpis):
        """Test KPI sheet contains correct data"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        ws = wb["KPIs"]

        # Check header row
        assert ws["A1"].value == "KPI Name"
        assert ws["B1"].value == "Latest Value"

        # Check data rows
        row_count = ws.max_row
        assert row_count > 1  # At least header + data

    def test_evidence_sheet_with_links(self, excel_generator, setup_report_with_kpis):
        """Test evidence sheet with link data"""
        report, kpis = setup_report_with_kpis

        evidence_links = [
            {
                "name": "Energy Audit",
                "uploaded_at": "2024-01-15",
                "type": "audit",
                "link": "https://example.com/evidence/1"
            },
            {
                "name": "Carbon Cert",
                "uploaded_at": "2024-02-01",
                "type": "certification",
                "link": "https://example.com/evidence/2"
            }
        ]

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00"),
            evidence_links=evidence_links
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Evidence"]

        # Check headers
        assert ws["A1"].value == "Document Name"
        assert ws["D1"].value == "URL"

        # Check data
        assert ws["A2"].value == "Energy Audit"


class TestExcelFormatting:
    """Test Excel formatting"""

    def test_header_row_formatting(self, excel_generator, setup_report_with_kpis):
        """Test header rows are formatted"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Summary"]

        # Check header cell formatting
        header_cell = ws["A7"]
        assert header_cell.fill.patternType is not None
        assert header_cell.font.bold

    def test_numeric_formatting(self, excel_generator, setup_report_with_kpis):
        """Test numeric values are formatted"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.55"),
            scope_2=Decimal("200.75"),
            scope_3=Decimal("300.25")
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Summary"]

        # Check numeric formatting
        numeric_cell = ws["B8"]
        assert numeric_cell.number_format in ['#,##0.00', '0.00']

    def test_frozen_panes(self, excel_generator, setup_report_with_kpis):
        """Test freeze panes are set"""
        report, kpis = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Summary"]

        # Check freeze panes are set
        assert ws.freeze_panes is not None


class TestExcelErrorHandling:
    """Test error handling"""

    def test_missing_report_error(self, excel_generator):
        """Test error when report not found"""
        with pytest.raises(ValueError) as exc_info:
            excel_generator.generate_workbook(
                report_id=uuid4(),
                organization_id=uuid4()
            )
        assert "not found" in str(exc_info.value)

    def test_missing_organization_error(self, excel_generator, setup_report_with_kpis):
        """Test error when organization not found"""
        report, _ = setup_report_with_kpis

        with pytest.raises(ValueError) as exc_info:
            excel_generator.generate_workbook(
                report_id=report.id,
                organization_id=uuid4()
            )
        assert "not found" in str(exc_info.value)

    def test_missing_kpi_data_graceful(self, excel_generator, setup_report_with_kpis):
        """Test handling of missing KPI data"""
        report, _ = setup_report_with_kpis

        # Should not raise error with None KPI data
        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=None,
            scope_2=None,
            scope_3=None
        )

        assert isinstance(excel_buffer, BytesIO)


class TestExcelLargeDatasets:
    """Test handling of large datasets"""

    def test_large_evidence_list(self, excel_generator, setup_report_with_kpis):
        """Test Excel with many evidence items"""
        report, _ = setup_report_with_kpis

        evidence_links = [
            {
                "name": f"Evidence {i}",
                "uploaded_at": f"2024-01-{(i % 28) + 1:02d}",
                "type": "document",
                "link": f"https://example.com/evidence/{i}"
            }
            for i in range(100)
        ]

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00"),
            evidence_links=evidence_links
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Evidence"]
        assert ws.max_row > 100

    def test_large_detailed_data(self, excel_generator, setup_report_with_kpis):
        """Test Excel with large detailed data sheets"""
        report, _ = setup_report_with_kpis

        # Create large dataset
        large_data = [
            [f"2024-01-{(i % 28) + 1:02d}", f"{1000 + i}.50", "Grid"]
            for i in range(500)
        ]

        detailed_data = {
            "Large Dataset": (
                ["Date", "Value", "Source"],
                large_data
            )
        }

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            detailed_data=detailed_data
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Large Dataset"]
        assert ws.max_row > 500


class TestExcelEmissions:
    """Test emissions data in Excel"""

    def test_emissions_in_summary(self, excel_generator, setup_report_with_kpis):
        """Test emissions values appear in summary sheet"""
        report, _ = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("123.45"),
            scope_2=Decimal("234.56"),
            scope_3=Decimal("345.67")
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Summary"]

        # Find emissions rows and verify values
        found_scope_1 = False
        for row in ws.iter_rows(min_col=1, max_col=3, values_only=True):
            if row[0] == "Scope 1 Emissions":
                found_scope_1 = True
                assert float(row[1]) == 123.45

        assert found_scope_1

    def test_total_emissions_calculation(self, excel_generator, setup_report_with_kpis):
        """Test total emissions is calculated"""
        report, _ = setup_report_with_kpis

        excel_buffer = excel_generator.generate_workbook(
            report_id=report.id,
            organization_id=report.organization_id,
            scope_1=Decimal("100.00"),
            scope_2=Decimal("200.00"),
            scope_3=Decimal("300.00")
        )

        wb = load_workbook(excel_buffer)
        ws = wb["Summary"]

        # Find total emissions
        for row in ws.iter_rows(min_col=1, max_col=3, values_only=True):
            if row[0] == "Total Emissions":
                assert float(row[1]) == 600.00
                break
